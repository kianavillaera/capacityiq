"""Attendance report building and styled Excel export."""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

ROW_FILL = {
    "Ghost": PatternFill("solid", fgColor="FFCCCC"),
    "Incomplete": PatternFill("solid", fgColor="FFF2CC"),
    "Compliant": PatternFill("solid", fgColor="CCFFDD"),
}
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
ONCALL_FILL = PatternFill("solid", fgColor="FFD966")

SHEET_FILL_WEEKLY = {
    "ghost_resources": PatternFill("solid", fgColor="FFCCCC"),
    "incomplete_loggers": PatternFill("solid", fgColor="FFF2CC"),
    "orphan_tc_users": PatternFill("solid", fgColor="E2EFDA"),
}
ROSTER_META = ["Name", "Pod", "Technology", "Specialisation", "Seniority", "Location", "Email"]


def _month_slug(month_label: str) -> str:
    """Derive a safe Excel sheet-name slug from a month label.

    Examples: 'Jun 2026' → 'jun', 'Jul 2026' → 'jul', 'July 2026' → 'july'.
    """
    import re
    return re.sub(r"[^\w]", "_", month_label.strip().split()[0]).lower().strip("_")


def _is_exception(name: str, exceptions: list) -> bool:
    """Return True if this roster name belongs to a partial-hours exception.

    Matches against the first-name part (after the comma in 'Surname, Firstname').
    The check is case-insensitive and bidirectional so both 'Mark' and 'Mark Anthony'
    will match an exception entry of 'Mark Anthony'.
    """
    if not exceptions or not name:
        return False
    parts = str(name).split(",", 1)
    first = (parts[1] if len(parts) > 1 else parts[0]).strip().lower()
    return any(
        exc.strip().lower() == first
        or exc.strip().lower() in first
        or first in exc.strip().lower()
        for exc in exceptions
    )


def build_weekly_attendance(
    tc_week: pd.DataFrame,
    tc_oncall_week: pd.DataFrame,
    resources_matched: pd.DataFrame,
    week: pd.Timestamp,
    hours_threshold: int = 40,
    partial_hours_exceptions: Optional[list] = None,
) -> tuple:
    """Build attendance stats for a single week. Returns (roster, orphans, ghost, incomplete, full, day_cols)."""
    week_dates = [week + pd.Timedelta(days=i) for i in range(7)]
    day_cols = [d.strftime("%a %d/%m") for d in week_dates]

    daily = (
        tc_week.assign(day_label=tc_week["Date"].dt.strftime("%a %d/%m"))
        .pivot_table(index="_uid", columns="day_label", values="Time worked", aggfunc="sum")
        .reindex(columns=day_cols)
        .fillna(0)
        .round(2)
        .reset_index()
    )

    per_uid = (
        tc_week.groupby("_uid")
        .agg(
            name_tc=("User", "first"),
            hours_logged=("Time worked", "sum"),
            days_logged=("Date", "nunique"),
            categories=("Category", lambda x: " | ".join(sorted(set(x.dropna())))),
        )
        .reset_index()
        .assign(hours_logged=lambda d: d["hours_logged"].round(2))
        .merge(daily, on="_uid", how="left")
    )

    task_hrs = (
        tc_week[~tc_week["is_gen"] & (tc_week["Category"] == "Task work")]
        .groupby("_uid")["Time worked"].sum().round(2).rename("hours_task")
    )
    gen_hrs = tc_week[tc_week["is_gen"]].groupby("_uid")["Time worked"].sum().round(2).rename("hours_gen")
    other_hrs = (
        tc_week[~tc_week["is_gen"] & (tc_week["Category"] != "Task work")]
        .groupby("_uid")["Time worked"].sum().round(2).rename("hours_other")
    )
    oncall_hrs = tc_oncall_week.groupby("_uid")["Time worked"].sum().round(2).rename("hours_oncall")

    per_uid = (
        per_uid
        .join(task_hrs, on="_uid")
        .join(gen_hrs, on="_uid")
        .join(other_hrs, on="_uid")
        .join(oncall_hrs, on="_uid")
        .fillna({"hours_task": 0, "hours_gen": 0, "hours_other": 0, "hours_oncall": 0})
    )

    roster = (
        resources_matched[ROSTER_META + ["tc_uid"]]
        .merge(per_uid.rename(columns={"_uid": "tc_uid"}), on="tc_uid", how="left")
        .drop(columns="tc_uid")
    )
    for c in ["hours_logged", "hours_task", "hours_gen", "hours_other", "hours_oncall"] + day_cols:
        roster[c] = roster[c].fillna(0)

    orphans = per_uid[~per_uid["_uid"].isin(resources_matched["tc_uid"].dropna())]

    ghost = roster[roster["hours_logged"] == 0][
        ROSTER_META + ["hours_oncall"]
    ].reset_index(drop=True)

    _exc_mask = roster["Name"].apply(lambda n: _is_exception(n, partial_hours_exceptions or []))
    incomplete = (
        roster[(roster["hours_logged"] > 0) & (roster["hours_logged"] < hours_threshold) & ~_exc_mask]
        .copy()
        .assign(shortfall=lambda d: (hours_threshold - d["hours_logged"]).round(2))
        .sort_values("shortfall", ascending=False)
        .reset_index(drop=True)
    )

    full = _build_full_roster(roster, hours_threshold, day_cols, partial_hours_exceptions)

    logger.info(
        "Week %s: %d roster | compliant=%d incomplete=%d ghost=%d orphans=%d",
        week.date(),
        len(roster),
        int((roster["hours_logged"] >= hours_threshold).sum()),
        len(incomplete),
        len(ghost),
        len(orphans),
    )

    return roster, orphans, ghost, incomplete, full, day_cols

def build_monthly_attendance(
    tc: pd.DataFrame,
    tc_oncall: pd.DataFrame,
    resources_matched: pd.DataFrame,
    weeks: list,
    month_threshold: int = 200,
    hours_threshold: int = 40,
    partial_hours_exceptions: Optional[list] = None,
) -> tuple:
    """Build monthly attendance stats (all weeks combined). Returns (roster, orphans, ghost, incomplete, full, wk_cols)."""
    wk_cols = []
    week_series = {}
    for w in weeks:
        col = "Wk_" + w.strftime("%d%b")
        wk_cols.append(col)
        week_series[col] = (
            tc[tc["week_start"] == w].groupby("_uid")["Time worked"]
            .sum().round(2).rename(col)
        )

    per_uid = (
        tc.groupby("_uid")
        .agg(
            name_tc=("User", "first"),
            hours_logged=("Time worked", "sum"),
            days_logged=("Date", "nunique"),
            categories=("Category", lambda x: " | ".join(sorted(set(x.dropna())))),
        )
        .reset_index()
        .assign(hours_logged=lambda d: d["hours_logged"].round(2))
    )
    for col, s in week_series.items():
        per_uid = per_uid.join(s, on="_uid")
    per_uid[wk_cols] = per_uid[wk_cols].fillna(0)

    task_hrs = (
        tc[~tc["is_gen"] & (tc["Category"] == "Task work")]
        .groupby("_uid")["Time worked"].sum().round(2).rename("hours_task")
    )
    gen_hrs = tc[tc["is_gen"]].groupby("_uid")["Time worked"].sum().round(2).rename("hours_gen")
    other_hrs = (
        tc[~tc["is_gen"] & (tc["Category"] != "Task work")]
        .groupby("_uid")["Time worked"].sum().round(2).rename("hours_other")
    )
    oncall_hrs = tc_oncall.groupby("_uid")["Time worked"].sum().round(2).rename("hours_oncall")

    per_uid = (
        per_uid
        .join(task_hrs, on="_uid")
        .join(gen_hrs, on="_uid")
        .join(other_hrs, on="_uid")
        .join(oncall_hrs, on="_uid")
        .fillna({"hours_task": 0, "hours_gen": 0, "hours_other": 0, "hours_oncall": 0})
    )

    roster = (
        resources_matched[ROSTER_META + ["tc_uid"]]
        .merge(per_uid.rename(columns={"_uid": "tc_uid"}), on="tc_uid", how="left")
        .drop(columns="tc_uid")
    )
    for c in ["hours_logged", "hours_task", "hours_gen", "hours_other", "hours_oncall", "days_logged"] + wk_cols:
        roster[c] = roster[c].fillna(0)
    roster["categories"] = roster["categories"].fillna("")

    # How many of the N weeks did each person meet the weekly threshold?
    if wk_cols:
        roster["weeks_compliant"] = (roster[wk_cols] >= hours_threshold).sum(axis=1)
        roster["weeks_compliant"] = (
            roster["weeks_compliant"].astype(str) + "/" + str(len(wk_cols))
        )

    orphans = per_uid[~per_uid["_uid"].isin(resources_matched["tc_uid"].dropna())]

    ghost = roster[roster["hours_logged"] == 0][
        ROSTER_META + ["hours_oncall"]
    ].reset_index(drop=True)

    _exc_mask = roster["Name"].apply(lambda n: _is_exception(n, partial_hours_exceptions or []))
    incomplete = (
        roster[(roster["hours_logged"] > 0) & (roster["hours_logged"] < month_threshold) & ~_exc_mask]
        .copy()
        .assign(shortfall=lambda d: (month_threshold - d["hours_logged"]).round(2))
        .sort_values("shortfall", ascending=False)
        .reset_index(drop=True)
    )

    full = _build_full_roster(roster, month_threshold, wk_cols, partial_hours_exceptions)

    logger.info(
        "Monthly attendance: %d roster | compliant=%d incomplete=%d ghost=%d orphans=%d",
        len(roster),
        int((roster["hours_logged"] >= month_threshold).sum()),
        len(incomplete),
        len(ghost),
        len(orphans),
    )

    return roster, orphans, ghost, incomplete, full, wk_cols

def _build_full_roster(
    roster: pd.DataFrame,
    threshold: int,
    time_cols: list,
    partial_hours_exceptions: Optional[list] = None,
) -> pd.DataFrame:
    """Build the full sorted roster DataFrame with status and shortfall columns."""
    base_cols = [c for c in ROSTER_META if c in roster.columns]
    hour_cols = [
        c for c in ["hours_logged", "hours_task", "hours_gen", "hours_other",
                     "hours_oncall", "shortfall", "days_logged", "weeks_compliant",
                     "categories"]
        if c in roster.columns or c == "shortfall"
    ]
    _exc = partial_hours_exceptions or []

    def _status(row) -> str:
        h = row["hours_logged"]
        if _exc and _is_exception(str(row.get("Name", "")), _exc):
            return "Compliant" if h > 0 else "Ghost"
        return "Compliant" if h >= threshold else ("Ghost" if h == 0 else "Incomplete")

    def _shortfall(row) -> float:
        h = row["hours_logged"]
        if _exc and _is_exception(str(row.get("Name", "")), _exc) and h > 0:
            return 0.0
        return max(0.0, round(float(threshold - h), 2))

    return (
        roster.assign(
            status=lambda d: d.apply(_status, axis=1),
            shortfall=lambda d: d.apply(_shortfall, axis=1),
        )[[*base_cols, "status", "hours_logged", "hours_task", "hours_gen",
           "hours_other", "hours_oncall", "shortfall", "days_logged",
           *(["weeks_compliant"] if "weeks_compliant" in roster.columns else []),
           "categories"] + time_cols]
        .assign(_order=lambda d: d["status"].map({"Compliant": 0, "Incomplete": 1, "Ghost": 2}))
        .sort_values(["_order", "hours_logged"])
        .drop(columns="_order")
        .reset_index(drop=True)
    )

def _style_sheet(ws, fill=None) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    if fill:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = fill

def _highlight_oncall_col(ws) -> None:
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "hours_oncall":
            for row_num in range(1, ws.max_row + 1):
                ws.cell(row=row_num, column=col_idx).fill = ONCALL_FILL
            break

def _style_full_roster(ws, status_col_idx: int) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fill = ROW_FILL.get(row[status_col_idx].value)
        if fill:
            for cell in row:
                cell.fill = fill
    _highlight_oncall_col(ws)

def _build_weekly_legend(ws, week_label: str, threshold: int) -> None:
    """Write the legend sheet for a weekly attendance workbook."""
    TITLE_FONT = Font(bold=True, size=12)
    SECTION_FONT = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    def _row(label, value, label_fill=None):
        r = ws.max_row + 1
        a, b = ws.cell(r, 1, label), ws.cell(r, 2, value)
        a.alignment = b.alignment = WRAP
        if label_fill:
            a.fill = label_fill

    ws.cell(1, 1, f"Legend - Attendance Analysis  ·  Week {week_label}").font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.append([])

    ws.append(["ROW COLOURS (full_roster sheet)", ""])
    ws["A3"].font = SECTION_FONT
    _row("Compliant  (green)", f"hours_logged ≥ {threshold} h", ROW_FILL["Compliant"])
    _row("Incomplete  (yellow)", f"0 < hours_logged < {threshold} h", ROW_FILL["Incomplete"])
    _row("Ghost  (red)", "hours_logged = 0 - nothing logged at all", ROW_FILL["Ghost"])
    ws.append([])

    ws.append(["AMBER COLUMN", ""])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    _row(
        "hours_oncall  (amber)",
        "On-call hours for the week. Excluded from the 40 h threshold - logged separately by the system.",
        ONCALL_FILL,
    )
    ws.append([])

    ws.append(["COLUMN GUIDE", ""])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    for col, desc in [
        ("status", "Compliant / Incomplete / Ghost"),
        ("hours_logged", "Total hours logged (task work + GEN + sick/holiday). Excludes on-call."),
        ("hours_task", "Task work hours only (Category = 'Task work', Task code not starting with GEN)."),
        ("hours_gen", "GEN overhead hours (Task code starts with GEN). Count toward hours_logged but are not billable."),
        ("hours_other", "Hours in other categories (Sick/Holiday, Available to work, etc.)."),
        ("hours_oncall", "On-call hours. Not counted in hours_logged or the 40 h threshold."),
        ("shortfall", f"{threshold} − hours_logged, clipped to 0 for compliant resources."),
        ("days_logged", "Number of distinct calendar days with at least one entry."),
        ("categories", "Pipe-separated list of distinct Category values logged."),
        ("Mon dd/mm … Sun dd/mm", "Hours logged on each day of the 7-day week span."),
    ]:
        _row(col, desc)

def write_weekly_attendance_report(
    output_path: Path,
    full: pd.DataFrame,
    ghost: pd.DataFrame,
    incomplete: pd.DataFrame,
    orphans: pd.DataFrame,
    day_cols: list,
    week: pd.Timestamp,
    hours_threshold: int = 40,
) -> None:
    """Write the weekly attendance workbook to disk."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    orphan_export_cols = [
        "_uid", "name_tc", "hours_logged", "hours_task",
        "hours_gen", "hours_other", "hours_oncall", "days_logged",
    ] + day_cols
    orphan_export_cols = [c for c in orphan_export_cols if c in orphans.columns]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        full.to_excel(writer, sheet_name="full_roster", index=False)
        ghost.to_excel(writer, sheet_name="ghost_resources", index=False)
        incomplete.to_excel(writer, sheet_name="incomplete_loggers", index=False)
        orphans[orphan_export_cols].to_excel(writer, sheet_name="orphan_tc_users", index=False)

        status_col_idx = list(full.columns).index("status")
        _style_full_roster(writer.sheets["full_roster"], status_col_idx)
        for sheet_name, fill in SHEET_FILL_WEEKLY.items():
            if sheet_name in writer.sheets:
                _style_sheet(writer.sheets[sheet_name], fill=fill)
                _highlight_oncall_col(writer.sheets[sheet_name])

        wb = writer.book
        ws_legend = wb.create_sheet("legend")
        _build_weekly_legend(ws_legend, str(week.date()), hours_threshold)

    logger.info("Weekly attendance report written: %s", output_path)

def write_monthly_attendance_report(
    output_path: Path,
    full_m: pd.DataFrame,
    ghost_m: pd.DataFrame,
    incomplete_m: pd.DataFrame,
    orphans_m: pd.DataFrame,
    wk_cols: list,
    week_rosters: dict,
    month_label: str,
    hours_threshold: int = 40,
    month_threshold: int = 200,
    sub_period_rosters: Optional[list] = None,  # list of (label, full_df, threshold)
) -> None:
    """Write the monthly attendance workbook including per-week sheets."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    orphan_base_cols = ["_uid", "name_tc", "hours_logged", "hours_task",
                        "hours_gen", "hours_other", "hours_oncall", "days_logged"]

    slug = _month_slug(month_label)
    sheet_full    = f"{slug}_full_roster"
    sheet_ghost   = f"ghost_{slug}"
    sheet_inc     = f"incomplete_{slug}"
    sheet_orphans = f"orphans_{slug}"
    sheet_fills   = {
        sheet_ghost:   PatternFill("solid", fgColor="FFCCCC"),
        sheet_inc:     PatternFill("solid", fgColor="FFF2CC"),
        sheet_orphans: PatternFill("solid", fgColor="E2EFDA"),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        full_m.to_excel(writer, sheet_name=sheet_full, index=False)
        ghost_m.to_excel(writer, sheet_name=sheet_ghost, index=False)
        incomplete_m.to_excel(writer, sheet_name=sheet_inc, index=False)

        orphan_cols = [c for c in orphan_base_cols if c in orphans_m.columns]
        orphans_m[orphan_cols].to_excel(writer, sheet_name=sheet_orphans, index=False)

        status_col_idx = list(full_m.columns).index("status")
        _style_full_roster(writer.sheets[sheet_full], status_col_idx)
        for sname, sfill in sheet_fills.items():
            if sname in writer.sheets:
                _style_sheet(writer.sheets[sname], fill=sfill)
                _highlight_oncall_col(writer.sheets[sname])

        # Per-week sheets
        for w, (roster_w, orphans_w, day_cols_w) in sorted(week_rosters.items()):
            sheet_label = f"week_{w.date()}"
            full_w = _build_full_roster(roster_w, hours_threshold, day_cols_w)
            full_w.to_excel(writer, sheet_name=sheet_label, index=False)
            _style_full_roster(
                writer.sheets[sheet_label], list(full_w.columns).index("status")
            )

        # Sub-period full-roster sheets (e.g. June-only, July-only)
        for sp_label, sp_full, _sp_thresh in (sub_period_rosters or []):
            sp_slug = _month_slug(sp_label)
            sp_sheet = f"{sp_slug}_full_roster"
            sp_full.to_excel(writer, sheet_name=sp_sheet, index=False)
            _style_full_roster(writer.sheets[sp_sheet], list(sp_full.columns).index("status"))

        wb = writer.book
        ws_legend = wb.create_sheet("legend")
        _build_monthly_legend(ws_legend, month_label, slug, hours_threshold, month_threshold, wk_cols, week_rosters, sub_period_rosters)

    logger.info("Monthly attendance report written: %s", output_path)

def _build_monthly_legend(ws, month_label, slug, week_threshold, month_threshold, wk_cols, week_rosters, sub_period_rosters=None) -> None:
    TITLE_FONT = Font(bold=True, size=12)
    SECTION_FONT = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70

    sheet_full    = f"{slug}_full_roster"
    sheet_ghost   = f"ghost_{slug}"
    sheet_inc     = f"incomplete_{slug}"
    sheet_orphans = f"orphans_{slug}"

    def _row(label, value, label_fill=None):
        r = ws.max_row + 1
        a, b = ws.cell(r, 1, label), ws.cell(r, 2, value)
        a.alignment = b.alignment = WRAP
        if label_fill:
            a.fill = label_fill

    ws.cell(1, 1, f"Legend - Compliance Analysis  ·  {month_label}").font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.append([])

    ws.append([f"ROW COLOURS ({sheet_full} sheet)", ""])
    ws["A3"].font = SECTION_FONT
    _row("Compliant (green)", f"hours_logged ≥ {month_threshold} h (whole month)", ROW_FILL["Compliant"])
    _row("Incomplete (yellow)", f"0 < hours_logged < {month_threshold} h", ROW_FILL["Incomplete"])
    _row("Ghost (red)", f"hours_logged = 0 for the entire {month_label} period", ROW_FILL["Ghost"])
    ws.append([])

    ws.append(["AMBER COLUMN", ""])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    _row("hours_oncall (amber)", "On-call hours. Excluded from the monthly threshold.", ONCALL_FILL)
    ws.append([])

    ws.append(["SHEETS", ""])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    sheets_desc = [
        (sheet_full,    f"All roster members for {month_label} sorted Compliant, Incomplete, Ghost."),
        (sheet_ghost,   "Roster members with 0 hours logged for the entire period."),
        (sheet_inc,     f"0 < hours_logged < {month_threshold} h, sorted by shortfall."),
        (sheet_orphans, "TC users who logged time but are not in the resource roster."),
    ]
    for sp_label, _sp_full, sp_thresh in (sub_period_rosters or []):
        sp_slug = _month_slug(sp_label)
        sheets_desc.append((
            f"{sp_slug}_full_roster",
            f"{sp_label} only — full roster coloured by compliance status (\u2265 {sp_thresh} h threshold for this period).",
        ))
    for w in sorted(week_rosters.keys()):
        sheets_desc.append((f"week_{w.date()}", f"Per-week detail for week starting {w.date()}."))
    sheets_desc.append(("legend", "This sheet."))
    for sheet, desc in sheets_desc:
        _row(sheet, desc)
